
from pathlib import Path
from typing import Any, Dict, List
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# Colour palette
HEADER_BG     = "1F3864"   # dark navy
HEADER_FG     = "FFFFFF"
PASS_BG       = "C6EFCE"   # light green
FAIL_BG       = "FFC7CE"   # light red
SKIP_BG       = "FFEB9C"   # light yellow
ALT_ROW_BG    = "EBF3FB"   # very light blue
TITLE_BG      = "2F75B6"   # medium blue

COLUMNS = [
    ("Test Case ID",    12),
    ("Module",          18),
    ("Title",           35),
    ("Type",            12),
    ("Priority",        10),
    ("Preconditions",   25),
    ("Steps",           50),
    ("Expected Result", 40),
    ("Test Data",       25),
    ("Status",          12),
    ("Actual Result",   40),
    ("Execution Time",  20),
    ("Error Message",   35),
    ("Created At",      20),
]


def _header_style() -> dict:
    return {
        "font": Font(bold=True, color=HEADER_FG, size=11, name="Calibri"),
        "fill": PatternFill("solid", fgColor=HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right=Side(style="thin", color="AAAAAA"),
        ),
    }


def _apply_style(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


class ExcelHandler:
    """Create and update Excel workbooks for test case management."""

    # ── Create ────────────────────────────────────────────────────────────
    def create(self, test_cases: List[Dict[str, Any]], path: Path):
        wb = openpyxl.Workbook()

        # ── Summary Sheet ──────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"
        self._build_summary(ws_sum, test_cases)

        # ── Test Cases Sheet ───────────────────────────────────────────
        ws_tc = wb.create_sheet("Test Cases")
        self._build_test_cases_sheet(ws_tc, test_cases)

        wb.save(path)

    def _build_summary(self, ws, test_cases: List[Dict]):
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # Title
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = "AI Test Automation Report"
        _apply_style(
            title_cell,
            font=Font(bold=True, size=16, color="FFFFFF", name="Calibri"),
            fill=PatternFill("solid", fgColor=TITLE_BG),
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        ws.row_dimensions[1].height = 40

        rows = [
            ("Generated At",   datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Cases",    len(test_cases)),
            ("Status",         "Pending Execution"),
            ("Modules",        ", ".join({tc["module"] for tc in test_cases})),
        ]
        for i, (label, val) in enumerate(rows, start=2):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = val
            ws[f"A{i}"].font = Font(bold=True, name="Calibri")

    def _build_test_cases_sheet(self, ws, test_cases: List[Dict]):
        # Freeze pane
        ws.freeze_panes = "A3"

        # Title row
        ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
        title = ws["A1"]
        title.value = "Test Case Repository — AI Generated"
        _apply_style(
            title,
            font=Font(bold=True, size=13, color="FFFFFF", name="Calibri"),
            fill=PatternFill("solid", fgColor=TITLE_BG),
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        ws.row_dimensions[1].height = 32

        # Header row
        hstyle = _header_style()
        for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            _apply_style(cell, **hstyle)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        ws.row_dimensions[2].height = 28

        # Data rows
        for row_idx, tc in enumerate(test_cases, start=3):
            steps_text = "\n".join(
                f"{i+1}. {s}" for i, s in enumerate(tc.get("steps", []))
            )
            test_data_raw = tc.get("test_data") or {}
            if isinstance(test_data_raw, dict):
                 test_data_text = "\n".join(f"{k}: {v}" for k, v in test_data_raw.items())
            elif isinstance(test_data_raw, list):
                 test_data_text = "\n".join(str(item) for item in test_data_raw)
            else:
                 test_data_text = str(test_data_raw) if test_data_raw else "" 
            row = [
                tc.get("test_case_id", ""),
                tc.get("module", ""),
                tc.get("title", ""),
                tc.get("type", ""),
                tc.get("priority", ""),
                tc.get("preconditions", ""),
                steps_text,
                tc.get("expected_result", ""),
                test_data_text,
                tc.get("status", ""),
                tc.get("actual_result", ""),
                tc.get("execution_time", ""),
                tc.get("error_message", ""),
                tc.get("screenshot", ""),
                tc.get("created_at", ""),
            ]
            fill_color = ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(name="Calibri", size=10)
            ws.row_dimensions[row_idx].height = max(
                30, min(15 * (steps_text.count("\n") + 1), 120)
            )

    # ── Update Results ────────────────────────────────────────────────────
    def update_results(self, path: Path, results: Dict[str, Dict]):
        """Update Status, Actual Result, Execution Time, and Error columns."""
        wb = openpyxl.load_workbook(path)
        ws = wb["Test Cases"]

        # Build column index map from header row (row 2)
        col_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=2, column=col).value
            if header:
                col_map[header] = col

        status_col = col_map.get("Status")
        actual_col = col_map.get("Actual Result")
        exec_col   = col_map.get("Execution Time")
        error_col  = col_map.get("Error Message")
        id_col     = col_map.get("Test Case ID")

        if not all([status_col, id_col]):
            print("Could not find required columns for update")
            wb.save(path)
            return

        for row in range(3, ws.max_row + 1):
            tc_id_cell = ws.cell(row=row, column=id_col)
            tc_id = str(tc_id_cell.value or "").strip()
            if not tc_id:
                continue

            # Try exact match first, then case-insensitive
            result = results.get(tc_id) or next(
                (v for k, v in results.items() if k.upper() == tc_id.upper()), None
            )
            if not result:
                continue

            status = result.get("status", "")
            if status_col:
                cell = ws.cell(row=row, column=status_col, value=status)
                if status == "PASS":
                    cell.fill = PatternFill("solid", fgColor=PASS_BG)
                    cell.font = Font(bold=True, color="375623", name="Calibri")
                elif status == "FAIL":
                    cell.fill = PatternFill("solid", fgColor=FAIL_BG)
                    cell.font = Font(bold=True, color="9C0006", name="Calibri")
                else:
                    cell.fill = PatternFill("solid", fgColor=SKIP_BG)
                    cell.font = Font(bold=True, name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if actual_col:
                ws.cell(row=row, column=actual_col, value=result.get("actual_result", status))

            if exec_col:
                ws.cell(row=row, column=exec_col, value=result.get("executed_at", ""))

            if error_col:
                ws.cell(row=row, column=error_col, value=result.get("error_message", ""))

        # Update summary sheet
        if "Summary" in wb.sheetnames:
            ws_sum = wb["Summary"]
            passed  = sum(1 for r in results.values() if r.get("status") == "PASS")
            failed  = sum(1 for r in results.values() if r.get("status") == "FAIL")
            skipped = sum(1 for r in results.values() if r.get("status") == "SKIP")
            summary_data = {
                "Status":    "Completed",
                "Passed":    passed,
                "Failed":    failed,
                "Skipped":   skipped,
                "Completed": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            # append rows after existing data
            start_row = ws_sum.max_row + 2
            for label, val in summary_data.items():
                ws_sum[f"A{start_row}"] = label
                ws_sum[f"B{start_row}"] = val
                ws_sum[f"A{start_row}"].font = Font(bold=True, name="Calibri")
                start_row += 1

        wb.save(path)
